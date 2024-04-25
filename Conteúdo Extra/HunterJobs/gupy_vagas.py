from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from datetime import datetime
import time
import openpyxl
from openpyxl.styles import Font

# Solicita o input do usuário para o termo de busca
termo_busca = input("Digite a vaga ou empresa que deseja buscar: ")

# Configura o driver do selenium
options = Options()
options.headless = False  # Altere para True para executar sem abrir a janela do navegador
driver = webdriver.Chrome(options=options)
driver.get(f"https://portal.gupy.io/job-search/term={termo_busca}")

# Define o tempo total para rolagem da página em segundos
scroll_time = 30
start_time = time.time()

# Rola a página para baixo e espera carregar
while True:
    # Verifica se o tempo total para rolagem foi alcançado
    if time.time() - start_time > scroll_time:
        break

    driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
    time.sleep(5)

    # Verifica se chegou ao final da página
    last_element = driver.find_elements('xpath', "//ul[@class='sc-90466136-0 djVYjM']/li[last()]")
    if last_element:
        break

# Analisa o conteúdo HTML da página com a biblioteca BeautifulSoup
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
jobs = soup.find('ul', {'class': 'sc-90466136-0 djVYjM'})

# Cria uma nova planilha Excel para armazenar as informações coletadas
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Vagas de emprego"
header = ['Titulo da Vaga', 'Empresa', 'Localidade', 'Tipo Contratação', 'Links da Vaga', 'Data Extração']
sheet.append(header)

# Configurando Detalhes da Planilha
bold_font = Font(bold=True)
for cell in sheet[1]:
    cell.font = bold_font

# Retornando a data/Hora atual
now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

# Itera sobre as vagas encontradas e armazena as informações relevantes na planilha Excel
linha = 2
for job in jobs.find_all('li'):
    titulo_vaga = job.find('h4').text.strip()
    empresa = job.find('p', {'class': 'sc-efBctP dpAAMR sc-812f417a-5 hBmurm'}).text.strip()
    localizacao = job.find('p', {'class': 'sc-efBctP dpAAMR sc-812f417a-4 sc-812f417a-6 dkBRQd kaqRPR'}).text.strip()
    tipo_contratacao = job.find('p', {'class': 'sc-efBctP dpAAMR sc-812f417a-4 dkBRQd'}).text.strip()
    links = [link['href'] for link in job.find_all('a', {'class': 'sc-812f417a-1 fijAgW'})]
    print(f"\n"
          f"Vaga: {titulo_vaga}\n"
          f"Empresa: {empresa}\n"
          f"Local: {localizacao}\n"
          f"Tipo Contrato: {tipo_contratacao}\n"
          f"Links das Vagas: {links}\n"
          f"Data de Coleta e Envio dos Dados: {now}")

    sheet.cell(row=linha, column=1).value = titulo_vaga
    sheet.cell(row=linha, column=2).value = empresa
    sheet.cell(row=linha, column=3).value = localizacao
    sheet.cell(row=linha, column=4).value = tipo_contratacao
    sheet.cell(row=linha, column=5).value = ", ".join(links)
    sheet.cell(row=linha, column=6).value = now

    linha += 1

# Salva a planilha Excel
workbook.save(f"{termo_busca}_vagas.xlsx")

# Fecha o driver do selenium
driver.quit()

print(f"As vagas de emprego para '{termo_busca}' foram salvas em uma planilha Excel.")
